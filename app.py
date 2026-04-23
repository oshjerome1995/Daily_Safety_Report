from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify
import os
from datetime import datetime
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from PIL import Image as PILImage
from io import BytesIO
from database import init_db, get_db

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'svg'}

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Auto initialize database table
init_db()


# ----------------------------
# Helpers
# ----------------------------
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def query_records(filters=None, sort_field=None, sort_order="ASC"):
    filters = filters or {}
    db = get_db()
    query = "SELECT * FROM records WHERE 1=1"
    params = []

    # Filtering
    for key, val in filters.items():
        if val:
            query += f" AND {key} LIKE ?"
            params.append(f"%{val}%")

    # Sorting
    if sort_field:
        query += f" ORDER BY {sort_field} {sort_order}"
    else:
        query += " ORDER BY id DESC"

    records = db.execute(query, params).fetchall()
    db.close()
    return [dict(r) for r in records]


# ----------------------------
# Routes
# ----------------------------
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/records')
def fetch_records():
    filters = {
        "category": request.args.get("category"),
        "status": request.args.get("status"),
        "osh_rule": request.args.get("osh_rule"),
        "area_department": request.args.get("area_department")
    }
    sort_field = request.args.get("sort_field")
    sort_order = request.args.get("sort_order", "ASC").upper()
    records = query_records(filters, sort_field, sort_order)
    return jsonify(records)


@app.route('/add', methods=['GET', 'POST'])
def add_record():
    if request.method == 'POST':
        db = get_db()
        file = request.files.get('picture')
        add_file = request.files.get('additional_picture')

        cursor = db.execute("INSERT INTO records (detailed_observation, picture_path, additional_picture_path, "
                            "area_department, checked_by, action_done, date, so_timestamp, area_in_charge, "
                            "category, status, osh_rule, safety_officer) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                            (
                                request.form['detailed_observation'][:300],
                                None,
                                None,
                                request.form['area_department'][:300],
                                request.form['checked_by'][:300],
                                request.form['action_done'][:300],
                                request.form['date'],
                                datetime.now().strftime("%B %d %Y %H:%M:%S"),
                                request.form['area_in_charge'][:300],
                                request.form['category'],
                                "" if request.form['category']=="Good Observation" else request.form['status'],
                                "" if request.form['category']=="Good Observation" else request.form['osh_rule'],
                                request.form['safety_officer'][:300]
                            ))
        record_id = cursor.lastrowid

        # Save files if uploaded
        if file and allowed_file(file.filename):
            filename = f"{record_id}_{secure_filename(file.filename)}"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            db.execute("UPDATE records SET picture_path=? WHERE id=?", (f"uploads/{filename}", record_id))

        if add_file and allowed_file(add_file.filename):
            filename = f"{record_id}_add_{secure_filename(add_file.filename)}"
            add_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            db.execute("UPDATE records SET additional_picture_path=? WHERE id=?", (f"uploads/{filename}", record_id))

        db.commit()
        db.close()
        return redirect(url_for('index'))
    return render_template('record_form.html', record=None)


@app.route('/edit/<int:record_id>', methods=['GET', 'POST'])
def edit_record(record_id):
    db = get_db()
    record = db.execute("SELECT * FROM records WHERE id=?", (record_id,)).fetchone()
    if not record:
        db.close()
        return "Record not found", 404

    if request.method == 'POST':
        file = request.files.get('picture')
        add_file = request.files.get('additional_picture')

        if file and allowed_file(file.filename):
            filename = f"{record_id}_{secure_filename(file.filename)}"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            db.execute("UPDATE records SET picture_path=? WHERE id=?", (f"uploads/{filename}", record_id))

        if add_file and allowed_file(add_file.filename):
            filename = f"{record_id}_add_{secure_filename(add_file.filename)}"
            add_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            db.execute("UPDATE records SET additional_picture_path=? WHERE id=?", (f"uploads/{filename}", record_id))

        db.execute("""UPDATE records SET
            detailed_observation=?, area_department=?, checked_by=?, action_done=?, date=?, area_in_charge=?,
            category=?, status=?, osh_rule=?, safety_officer=?
            WHERE id=?""",
            (
                request.form['detailed_observation'][:300],
                request.form['area_department'][:300],
                request.form['checked_by'][:300],
                request.form['action_done'][:300],
                request.form['date'],
                request.form['area_in_charge'][:300],
                request.form['category'],
                "" if request.form['category']=="Good Observation" else request.form['status'],
                "" if request.form['category']=="Good Observation" else request.form['osh_rule'],
                request.form['safety_officer'][:300],
                record_id
            ))
        db.commit()
        db.close()
        return redirect(url_for('index'))

    db.close()
    return render_template('record_form.html', record=dict(record))


@app.route('/delete/<int:record_id>', methods=['POST'])
def delete_record(record_id):
    db = get_db()
    db.execute("DELETE FROM records WHERE id=?", (record_id,))
    db.commit()
    db.close()
    return redirect(url_for('index'))


@app.route('/export')
def export_excel():
    filters = {
        "category": request.args.get("category"),
        "status": request.args.get("status"),
        "osh_rule": request.args.get("osh_rule"),
        "area_department": request.args.get("area_department")
    }
    sort_field = request.args.get("sort_field")
    sort_order = request.args.get("sort_order", "ASC")
    records = query_records(filters, sort_field, sort_order)

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Safety Report"

    headers = ["ID","Observation","Picture","Area/Dept","Checked By","Action Done",
               "Date","SO Timestamp","Area In-Charge","Category","Additional Picture",
               "Status","OSH Rule","Safety Officer"]
    ws.append(headers)
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    MAX_WIDTH, MAX_HEIGHT = 120, 80

    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['K'].width = 18

    def insert_image(ws, col_letter, row_idx, img_path):
        if img_path and os.path.exists(os.path.join(app.static_folder, img_path)):
            img = PILImage.open(os.path.join(app.static_folder, img_path))
            img.thumbnail((MAX_WIDTH, MAX_HEIGHT))
            final_img = PILImage.new("RGBA", (MAX_WIDTH, MAX_HEIGHT), (255,255,255,0))
            final_img.paste(img, ((MAX_WIDTH - img.width)//2, (MAX_HEIGHT - img.height)//2))
            temp = BytesIO()
            final_img.convert("RGB").save(temp, format="PNG")
            temp.seek(0)
            ws.add_image(XLImage(temp), f"{col_letter}{row_idx}")

    for row_idx, r in enumerate(records, start=2):
        ws.append([
            r['id'], r['detailed_observation'], "", r['area_department'], r['checked_by'], r['action_done'],
            r['date'], r['so_timestamp'], r['area_in_charge'], r['category'], "",
            r['status'], r['osh_rule'], r['safety_officer']
        ])
        for col_idx in range(1, 15):
            ws.cell(row=row_idx, column=col_idx).alignment = alignment
        ws.row_dimensions[row_idx].height = MAX_HEIGHT * 0.75
        insert_image(ws, "C", row_idx, r.get("picture_path"))
        insert_image(ws, "K", row_idx, r.get("additional_picture_path"))

    for col_idx in range(1, 15):
        col_letter = get_column_letter(col_idx)
        if col_letter not in ['C','K']:
            max_len = max((len(str(ws.cell(r, col_idx).value or "")) for r in range(1, ws.max_row+1)), default=0)
            ws.column_dimensions[col_letter].width = min(max_len + 5, 40)

    output_file = f"DailySafetyReport_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    wb.save(output_file)
    return send_file(output_file, as_attachment=True)


# ----------------------------
# Run app
# ----------------------------
if __name__ == '__main__':
    app.run(debug=True)